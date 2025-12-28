"""
Excel data loader for Global Forest Watch data.
Uses Polars for efficient data loading and caching.
"""
import logging
from pathlib import Path
from typing import Dict, Optional, List
import polars as pl

from nexus.config.settings import settings

logger = logging.getLogger(__name__)


class ExcelLoader:
    """Loads Global Forest Watch Excel data with caching."""
    
    REQUIRED_SHEETS = [
        "Country tree cover loss",
        "Country primary loss", 
        "Country carbon data"
    ]
    
    def __init__(self, excel_path: Optional[Path] = None):
        """
        Initialize Excel loader.
        
        Args:
            excel_path: Path to Excel file. If None, uses default from settings.
        """
        if excel_path is None:
            excel_path = settings.raw_data_path / settings.excel_file
            
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
            
        self.excel_path = excel_path
        self._cache: Dict[str, pl.DataFrame] = {}
        self._validate_excel_structure()
        
    def _validate_excel_structure(self):
        """Validate that Excel file has required sheets."""
        try:
            # Use openpyxl to check sheet names (no pandas needed)
            from openpyxl import load_workbook
            
            wb = load_workbook(self.excel_path, read_only=True, data_only=True)
            available_sheets = wb.sheetnames
            wb.close()
            
            missing_sheets = set(self.REQUIRED_SHEETS) - set(available_sheets)
            if missing_sheets:
                raise ValueError(
                    f"Missing required sheets: {missing_sheets}. "
                    f"Available sheets: {available_sheets}"
                )
                
            logger.info(f"Excel file validated. Found sheets: {available_sheets}")
            
        except Exception as e:
            logger.error(f"Failed to validate Excel structure: {e}")
            raise
            
    def load_sheet(self, sheet_name: str, use_cache: bool = True) -> pl.DataFrame:
        """Load and cache an Excel sheet."""
        if use_cache and sheet_name in self._cache:
            logger.debug(f"Using cached data for sheet: {sheet_name}")
            return self._cache[sheet_name]
            
        logger.info(f"Loading sheet: {sheet_name}")
        
        try:
            # Use Polars directly to read Excel
            df = pl.read_excel(
                source=self.excel_path,
                sheet_name=sheet_name,
                engine='openpyxl'
            )
            
            # Validate
            if df.is_empty():
                raise ValueError(f"Sheet {sheet_name} is empty")
                
            # Cache the result
            self._cache[sheet_name] = df
            
            logger.info(
                f"Loaded {len(df):,} rows × {len(df.columns)} columns "
                f"from sheet: {sheet_name}"
            )
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to load sheet '{sheet_name}': {e}")
            raise
            
    def load_tree_cover_loss(self) -> pl.DataFrame:
        """Load tree cover loss data."""
        return self.load_sheet("Country tree cover loss")
        
    def load_primary_forest(self) -> pl.DataFrame:
        """Load primary forest loss data."""
        return self.load_sheet("Country primary loss")
        
    def load_carbon_data(self) -> pl.DataFrame:
        """Load carbon emissions data."""
        return self.load_sheet("Country carbon data")
        
    def load_all_sheets(self) -> Dict[str, pl.DataFrame]:
        """
        Load all required sheets.
        
        Returns:
            Dictionary mapping sheet names to DataFrames
        """
        result = {}
        for sheet_name in self.REQUIRED_SHEETS:
            result[sheet_name] = self.load_sheet(sheet_name)
        return result
        
    def get_sheet_info(self, sheet_name: str) -> Dict:
        """
        Get information about a sheet without fully loading it.
        
        Returns:
            Dictionary with sheet metadata
        """
        df = self.load_sheet(sheet_name)
        
        # Identify year columns
        year_columns = [
            col for col in df.columns 
            if any(year_str in col for year_str in [str(y) for y in range(2000, 2026)])
        ]
        
        # Identify static columns
        static_columns = [col for col in df.columns if col not in year_columns]
        
        return {
            "rows": len(df),
            "columns": len(df.columns),
            "year_columns": len(year_columns),
            "static_columns": len(static_columns),
            "years_range": self._extract_year_range(year_columns),
            "column_names": df.columns[:10],  # First 10 columns as sample
            "null_counts": {col: df[col].null_count() for col in df.columns[:5]},
        }
        
    def _extract_year_range(self, year_columns: List[str]) -> Optional[tuple]:
        """Extract min and max years from column names."""
        years = []
        for col in year_columns:
            # Extract 4-digit years from column names
            import re
            year_match = re.search(r'\d{4}', col)
            if year_match:
                years.append(int(year_match.group()))
                
        if years:
            return (min(years), max(years))
        return None
        
    def clear_cache(self):
        """Clear the cache."""
        self._cache.clear()
        logger.info("Cache cleared")
        
    def get_cache_size(self) -> int:
        """Get number of cached sheets."""
        return len(self._cache)


class DataValidator:
    """Validate loaded data for completeness and consistency."""
    
    @staticmethod
    def validate_columns(df: pl.DataFrame, required_columns: List[str]) -> bool:
        """
        Validate that DataFrame has required columns.
        
        Args:
            df: DataFrame to validate
            required_columns: List of required column names
            
        Returns:
            True if all required columns present
        """
        missing = set(required_columns) - set(df.columns)
        if missing:
            logger.error(f"Missing required columns: {missing}")
            return False
        return True
        
    @staticmethod
    def validate_year_columns(df: pl.DataFrame, start_year: int, end_year: int) -> bool:
        """
        Validate that DataFrame has expected year columns.
        
        Args:
            df: DataFrame to validate
            start_year: Expected start year
            end_year: Expected end year
            
        Returns:
            True if year columns are complete
        """
        expected_years = set(range(start_year, end_year + 1))
        
        # Extract years from column names
        year_columns = []
        for col in df.columns:
            import re
            year_match = re.search(r'(\d{4})', col)
            if year_match:
                year_columns.append(int(year_match.group()))
                
        found_years = set(year_columns)
        missing_years = expected_years - found_years
        
        if missing_years:
            logger.warning(f"Missing year columns for years: {sorted(missing_years)}")
            return False
            
        return True
        
    @staticmethod
    def validate_data_types(df: pl.DataFrame) -> bool:
        """
        Validate that numeric columns have appropriate data types.
        
        Returns:
            True if data types are valid
        """
        numeric_patterns = ['loss_ha', 'extent', 'area', 'carbon', 'emissions']
        
        for col in df.columns:
            if any(pattern in col.lower() for pattern in numeric_patterns):
                dtype = str(df[col].dtype)
                if 'Float' not in dtype and 'Int' not in dtype:
                    logger.warning(f"Column {col} should be numeric but is {dtype}")
                    return False
                    
        return True
        
    @staticmethod
    def check_data_completeness(df: pl.DataFrame, threshold: float = 0.7) -> float:
        """
        Check data completeness (percentage of non-null values).
        
        Args:
            df: DataFrame to check
            threshold: Minimum acceptable completeness ratio
            
        Returns:
            Completeness score (0-1)
        """
        total_cells = len(df) * len(df.columns)
        if total_cells == 0:
            return 0.0
            
        null_cells = sum(df[col].null_count() for col in df.columns)
        completeness = 1 - (null_cells / total_cells)
        
        if completeness < threshold:
            logger.warning(
                f"Data completeness {completeness:.1%} is below threshold {threshold:.1%}"
            )
            
        return completeness